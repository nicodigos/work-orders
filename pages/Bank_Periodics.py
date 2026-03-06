import io
import os
from pathlib import Path

import folium
import pandas as pd
import plotly.express as px
import streamlit as st
from folium.features import DivIcon
from folium.plugins import MarkerCluster
from openpyxl import load_workbook
from openpyxl.styles import Font
from streamlit_folium import st_folium
from utils.ms_graph_excel import (
    download_sharepoint_file_bytes,
    get_token_silent_or_raise,
    require_graph_login,
    resolve_drive_id,
    upload_sharepoint_file_bytes,
    write_temp_file,
)

# ==========================================
# PAGE CONFIG
# ==========================================
st.set_page_config(page_title="Banks Periodics", layout="wide")
st.title("Banks Periodics")
require_graph_login()

# ==========================================
# ENV (SharePoint file path + refresh cadence)
# ==========================================
BANKS_SP_PATH = os.getenv(
    "SP_BANKS_FILE_PATH",
    "General/9359-6633 QUEBEC INC/BGIS/Banks Periodics/2026.xlsx",
)

BANKS_REFRESH_SECONDS = 24 * 60 * 60  # 24 hours


@st.cache_data(show_spinner=False, ttl=BANKS_REFRESH_SECONDS)
def download_banks_excel_cached(sp_relative_path: str) -> str:
    token = get_token_silent_or_raise(
        "Not authenticated. Please connect in the main app (app.py).",
        "Session expired. Please reconnect in the main app (app.py).",
    )
    drive_id = resolve_drive_id(token)
    content = download_sharepoint_file_bytes(sp_relative_path, token, drive_id=drive_id)
    return write_temp_file(Path(sp_relative_path).name, content)


# ==========================================
# EXCEL HELPERS (VISIBLE SHEETS + HEADER DETECT)
# ==========================================
@st.cache_data(show_spinner=False)
def get_visible_sheet_names(xlsx_path: str) -> list[str]:
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    wb.close()
    return visible


def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    s = str(v).strip()
    return s == "" or s.lower() in {"nan", "none"}


@st.cache_data(show_spinner=False)
def detect_header_row(xlsx_path: str, sheet_name: str, scan_rows: int = 80) -> int:
    preview = pd.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        header=None,
        nrows=scan_rows,
        engine="openpyxl",
    )

    best_i = 0
    best_score = -1.0
    for i in range(len(preview)):
        row = preview.iloc[i].tolist()
        non_blank = [v for v in row if not _is_blank(v)]
        if len(non_blank) < 2:
            continue

        as_str = [str(v).strip() for v in non_blank]
        str_like = sum(1 for v in non_blank if isinstance(v, str))
        uniqueness = len(set(as_str)) / max(1, len(as_str))
        score = (len(non_blank) * 1.5) + (str_like * 2.0) + (uniqueness * 3.0)

        if score > best_score:
            best_score = score
            best_i = i

    return int(best_i)


@st.cache_data(show_spinner=False)
def read_sheet_with_detected_header(xlsx_path: str, sheet_name: str, header_row: int) -> pd.DataFrame:
    df = pd.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        header=header_row,
        engine="openpyxl",
    )
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ==========================================
# FIXED COLUMNS (BANK + ADDRESS)
# ==========================================
BANK_FALLBACKS = ["bank", "banco"]
ADDRESS_FALLBACKS = ["address", "adresse", "direccion", "direccion", "addr"]


def find_required_col(df: pd.DataFrame, fallbacks: list[str]) -> str | None:
    for c in df.columns:
        cl = str(c).strip().lower()
        if any(k in cl for k in fallbacks):
            return c
    return None


def to_text_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().replace({"": None, "nan": None, "None": None})


# ==========================================
# DONE / PENDING NORMALIZATION (CELL LEVEL)
# ==========================================
DONE_WORDS = {"done", "completed", "complete", "ok", "yes"}
PENDING_WORDS = {"pending", "pendiente", "to do", "todo", "open", "in progress"}
NOT_SCHEDULED_WORDS = {"not scheduled", "not schedule", "n/a", "na", "tbd"}


def normalize_status_cell(v) -> str | None:
    if _is_blank(v):
        return None
    s = str(v).strip().lower()
    if any(w in s for w in DONE_WORDS):
        return "Done"
    if any(w in s for w in PENDING_WORDS):
        return "Pending"
    if any(w in s for w in NOT_SCHEDULED_WORDS):
        return "Not Scheduled"
    return None


# ==========================================
# TABLE STYLING
# ==========================================
BANK_STYLES = {
    "TD": {"bg": "#54B848", "fg": "white"},
    "CIBC": {"bg": "#6f1729", "fg": "white"},
    "NB": {"bg": "white", "fg": "red"},
    "RBC": {"bg": "yellow", "fg": "blue"},
    "BMO": {"bg": "blue", "fg": "white"},
}
ADDRESS_STYLE = "background-color:#2b2b2b; color:white; font-weight:600;"


def _color_to_rgb(color: str) -> tuple[int, int, int]:
    c = str(color or "").strip().lower()
    named = {
        "white": (255, 255, 255),
        "black": (0, 0, 0),
        "red": (255, 0, 0),
        "blue": (0, 0, 255),
        "yellow": (255, 255, 0),
        "green": (0, 128, 0),
    }
    if c in named:
        return named[c]
    if c.startswith("#"):
        h = c[1:]
        if len(h) == 3:
            h = "".join(ch * 2 for ch in h)
        if len(h) == 6:
            try:
                return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
            except ValueError:
                pass
    return (30, 136, 229)


def _text_color_for_bg(bg_color: str) -> str:
    r, g, b = _color_to_rgb(bg_color)
    # Perceived brightness (WCAG-style weighting).
    brightness = (0.299 * r) + (0.587 * g) + (0.114 * b)
    return "black" if brightness >= 150 else "white"


def cell_style(v, is_bank_col: bool = False, is_addr_col: bool = False) -> str:
    if is_addr_col:
        return ADDRESS_STYLE

    if is_bank_col:
        if v in BANK_STYLES:
            s = BANK_STYLES[v]
            return f"background-color:{s['bg']}; color:{s['fg']}; font-weight:700;"
        return ""

    norm = normalize_status_cell(v)
    if norm == "Pending":
        return "background-color:#ffcdd2; color:#b71c1c; font-weight:600;"
    if norm == "Done":
        return "background-color:#c8e6c9; color:#1b5e20; font-weight:600;"
    return ""


def style_table(df: pd.DataFrame, bank_col: str, addr_col: str):
    def _row_style(row):
        styles = []
        for c in df.columns:
            styles.append(
                cell_style(
                    row.get(c),
                    is_bank_col=(c == bank_col),
                    is_addr_col=(c == addr_col),
                )
            )
        return styles

    return df.style.apply(_row_style, axis=1)


# ==========================================
# BAR CHART (DONE vs PENDING by column)
# ==========================================
def done_pending_by_column_barchart(df: pd.DataFrame, task_cols: list[str]):
    if df.empty or not task_cols:
        st.info("No data to summarize.")
        return

    rows = []
    for c in task_cols:
        norm = df[c].map(normalize_status_cell)
        done = int((norm == "Done").sum())
        pending = int((norm == "Pending").sum())
        total = done + pending
        if total == 0:
            continue

        rows.append({"Column": c, "Status": "Done", "Count": done, "Pct": done / total})
        rows.append({"Column": c, "Status": "Pending", "Count": pending, "Pct": pending / total})

    if not rows:
        st.info("No Done/Pending values found in the current selection.")
        return

    g = pd.DataFrame(rows)
    g["Label"] = (g["Pct"] * 100).round(0).astype(int).astype(str) + "%"

    col_order = [c for c in task_cols if c in g["Column"].unique()]
    g["Column"] = pd.Categorical(g["Column"], categories=col_order, ordered=True)

    fig = px.bar(
        g,
        x="Column",
        y="Count",
        color="Status",
        barmode="stack",
        text="Label",
        color_discrete_map={"Done": "#2e7d32", "Pending": "#c62828"},
        title="Completed by Work Type",
    )
    fig.update_traces(textposition="inside")
    fig.update_layout(
        xaxis_title="",
        yaxis_title="Count",
        legend_title_text="",
        margin=dict(l=20, r=20, t=60, b=40),
    )
    st.plotly_chart(fig, use_container_width=True)


def build_pending_long(
    df: pd.DataFrame,
    bank_col: str,
    addr_col: str,
    task_cols: list[str],
) -> pd.DataFrame:
    rows: list[dict] = []
    for _, row in df.iterrows():
        bank = row.get(bank_col)
        addr = row.get(addr_col)
        if _is_blank(bank) or _is_blank(addr):
            continue

        for service in task_cols:
            status = normalize_status_cell(row.get(service))
            if status == "Pending":
                rows.append(
                    {
                        "Bank": str(bank).strip(),
                        "Address": str(addr).strip(),
                        "Service": service,
                    }
                )
    return pd.DataFrame(rows)


def pending_matrix_by_bank_service(pending_long: pd.DataFrame) -> pd.DataFrame:
    if pending_long.empty:
        return pd.DataFrame()
    m = (
        pending_long.groupby(["Bank", "Service"])
        .size()
        .unstack(fill_value=0)
        .sort_index()
    )
    return m.transpose()


def style_pending_matrix(matrix: pd.DataFrame):
    df_show = matrix.reset_index().rename(columns={"index": "Service"})

    def _row_style(row):
        styles = []
        for c in df_show.columns:
            if c == "Service":
                styles.append("")
            else:
                v = row.get(c)
                if pd.notna(v) and float(v) != 0:
                    styles.append("background-color:#0d47a1; color:#ffffff; font-weight:700;")
                else:
                    styles.append("")
        return styles

    return df_show.style.apply(_row_style, axis=1)


def _norm_txt(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


MONTH_ORDER = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


def default_last_month_index(sheet_names: list[str]) -> int:
    if not sheet_names:
        return 0

    best_idx = len(sheet_names) - 1
    best_rank = -1
    for i, name in enumerate(sheet_names):
        rank = MONTH_ORDER.get(str(name).strip().lower(), -1)
        if rank >= best_rank:
            best_rank = rank
            best_idx = i
    return best_idx


def _norm_key(v) -> str:
    if _is_blank(v):
        return ""
    return str(v).strip().lower()


def city_from_address(addr) -> str:
    if _is_blank(addr):
        return "Unknown"
    parts = [p.strip() for p in str(addr).split(",") if p and str(p).strip()]
    if len(parts) >= 2:
        candidate = parts[-1]
        if any(ch.isdigit() for ch in candidate) and len(parts) > 1:
            candidate = parts[1]
        return candidate
    return "Unknown"


def load_masters_buildings_table(xlsx_path: str) -> pd.DataFrame:
    try:
        raw = pd.read_excel(xlsx_path, sheet_name="Masters", header=None, engine="openpyxl")
    except Exception:
        return pd.DataFrame()

    header_row = None
    col_map: dict[str, int] = {}
    key_aliases = {
        "bank": ["bank", "banco"],
        "address": ["address", "adresse", "direccion", "dirección", "addr"],
        "latitude": ["latitude", "latitud", "lat"],
        "longitude": ["longitude", "longitud", "lng", "lon"],
        "active": ["active", "activo"],
        "location google": ["location google", "google location", "google maps", "ubicacion google", "ubicación google"],
    }

    scan_rows = min(80, len(raw))
    for r in range(scan_rows):
        row_vals = raw.iloc[r].tolist()
        norm = [str(v).strip().lower() if not _is_blank(v) else "" for v in row_vals]
        tmp_map: dict[str, int] = {}
        for i, cell in enumerate(norm):
            if not cell:
                continue
            for canonical, aliases in key_aliases.items():
                if canonical in tmp_map:
                    continue
                if cell in aliases or any(a in cell for a in aliases):
                    tmp_map[canonical] = i
                    break

        if all(k in tmp_map for k in ["bank", "address", "latitude", "longitude"]):
            header_row = r
            col_map = tmp_map
            break

    if header_row is None:
        return pd.DataFrame()

    cols_to_take = ["bank", "address", "latitude", "longitude"] + [
        k for k in ["active", "location google"] if k in col_map
    ]
    data = raw.iloc[header_row + 1 :, [col_map[k] for k in cols_to_take]].copy()
    data.columns = [k.title() for k in cols_to_take]
    data = data.dropna(how="all")

    for c in ["Bank", "Address"]:
        if c in data.columns:
            data[c] = data[c].map(lambda x: None if _is_blank(x) else str(x).strip())

    if "Latitude" in data.columns:
        data["Latitude"] = pd.to_numeric(data["Latitude"], errors="coerce")
    if "Longitude" in data.columns:
        data["Longitude"] = pd.to_numeric(data["Longitude"], errors="coerce")

    data = data.dropna(subset=["Bank", "Address", "Latitude", "Longitude"])

    if "Active" in data.columns:
        def _to_bool(v):
            if isinstance(v, bool):
                return v
            s = _norm_key(v)
            if s in {"true", "1", "yes", "y", "si"}:
                return True
            if s in {"false", "0", "no", "n"}:
                return False
            return True
        data["Active"] = data["Active"].map(_to_bool)
        data = data[data["Active"]]

    return data.reset_index(drop=True)


def pending_counts_by_location_for_month(df_month: pd.DataFrame, bank_col: str, addr_col: str) -> pd.DataFrame:
    task_cols = [c for c in df_month.columns if c not in {bank_col, addr_col}]
    if not task_cols:
        return pd.DataFrame(columns=["Bank", "Address", "PendingCount"])

    d = df_month.copy()
    d["Bank"] = d[bank_col].map(lambda x: None if _is_blank(x) else str(x).strip())
    d["Address"] = d[addr_col].map(lambda x: None if _is_blank(x) else str(x).strip())
    d = d.dropna(subset=["Bank", "Address"])

    def _row_pending_count(row) -> int:
        count = 0
        for c in task_cols:
            if normalize_status_cell(row.get(c)) == "Pending":
                count += 1
        return count

    d["PendingCount"] = d.apply(_row_pending_count, axis=1)
    out = d.groupby(["Bank", "Address"], as_index=False)["PendingCount"].sum()
    return out.reset_index(drop=True)


def map_dataset_from_left_join_by_location(
    masters_df: pd.DataFrame,
    month_df: pd.DataFrame,
    bank_col: str,
    addr_col: str,
) -> pd.DataFrame:
    if masters_df.empty:
        return pd.DataFrame()

    month_loc = pending_counts_by_location_for_month(month_df, bank_col, addr_col)
    if month_loc.empty:
        return pd.DataFrame()
    month_loc["k_bank"] = month_loc["Bank"].map(_norm_key)
    month_loc["k_addr"] = month_loc["Address"].map(_norm_key)

    masters_loc = masters_df.copy()
    masters_loc["k_bank"] = masters_loc["Bank"].map(_norm_key)
    masters_loc["k_addr"] = masters_loc["Address"].map(_norm_key)
    masters_loc = masters_loc.drop_duplicates(subset=["k_bank", "k_addr"], keep="first")

    out = month_loc.merge(
        masters_loc[["k_bank", "k_addr", "Latitude", "Longitude"]],
        on=["k_bank", "k_addr"],
        how="left",
    )

    out["Latitude"] = pd.to_numeric(out["Latitude"], errors="coerce")
    out["Longitude"] = pd.to_numeric(out["Longitude"], errors="coerce")
    out = out.dropna(subset=["Latitude", "Longitude"])
    out["Label"] = out["PendingCount"].astype(str)
    out["Status"] = out["PendingCount"].map(lambda x: "Pending" if int(x) > 0 else "Not Pending")
    return out


def update_service_status_in_workbook(
    workbook_bytes: bytes,
    sheet_name: str,
    bank_col: str,
    addr_col: str,
    service_col: str,
    bank_value: str,
    address_value: str,
    new_status: str,
) -> bytes:
    with io.BytesIO(workbook_bytes) as bio:
        wb = load_workbook(filename=bio)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found.")

    ws = wb[sheet_name]
    tmp_path = write_temp_file("_banks_edit_source.xlsx", workbook_bytes)
    header_row_zero_based = detect_header_row(tmp_path, sheet_name)
    header_excel_row = header_row_zero_based + 1

    header_to_col_idx: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_excel_row, column=col_idx).value
        htxt = str(hv).strip() if hv is not None else ""
        if htxt and htxt not in header_to_col_idx:
            header_to_col_idx[htxt] = col_idx

    missing = [c for c in [bank_col, addr_col, service_col] if c not in header_to_col_idx]
    if missing:
        raise RuntimeError(f"Columns not found in worksheet header row: {missing}")

    bank_idx = header_to_col_idx[bank_col]
    addr_idx = header_to_col_idx[addr_col]
    service_idx = header_to_col_idx[service_col]

    target_row = None
    for r in range(header_excel_row + 1, ws.max_row + 1):
        bval = ws.cell(row=r, column=bank_idx).value
        aval = ws.cell(row=r, column=addr_idx).value
        if _norm_txt(bval) == _norm_txt(bank_value) and _norm_txt(aval) == _norm_txt(address_value):
            target_row = r
            break

    if target_row is None:
        raise RuntimeError("Could not locate the selected Bank + Address row in worksheet.")

    ws.cell(row=target_row, column=service_idx).value = new_status

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _unique_in_order(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for v in values:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def build_master_lists_from_workbook_bytes(workbook_bytes: bytes) -> tuple[list[tuple[str, str]], list[str]]:
    tmp_path = write_temp_file("_banks_master_source.xlsx", workbook_bytes)
    visible = get_visible_sheet_names(tmp_path)

    bank_addr_pairs: list[tuple[str, str]] = []
    services: list[str] = []

    for sheet in visible:
        header_row = detect_header_row(tmp_path, sheet)
        df = read_sheet_with_detected_header(tmp_path, sheet, header_row)
        if df.empty:
            continue

        bank_col = find_required_col(df, BANK_FALLBACKS)
        addr_col = find_required_col(df, ADDRESS_FALLBACKS)
        if not bank_col or not addr_col:
            continue

        df[bank_col] = to_text_series(df[bank_col])
        df[addr_col] = to_text_series(df[addr_col])
        sub = df[[bank_col, addr_col]].dropna()
        for _, row in sub.iterrows():
            b = str(row[bank_col]).strip()
            a = str(row[addr_col]).strip()
            if b and a:
                bank_addr_pairs.append((b, a))

        services.extend([c for c in df.columns if c not in {bank_col, addr_col}])

    bank_addr_pairs = _unique_in_order([f"{b}|||{a}" for b, a in bank_addr_pairs])
    bank_addr_final = [tuple(x.split("|||", 1)) for x in bank_addr_pairs]
    services_final = _unique_in_order([str(s).strip() for s in services if str(s).strip()])
    return bank_addr_final, services_final


def create_or_replace_masters_sheet(workbook_bytes: bytes, sheet_name: str = "Masters") -> bytes:
    banks, services = build_master_lists_from_workbook_bytes(workbook_bytes)

    with io.BytesIO(workbook_bytes) as bio:
        wb = load_workbook(filename=bio)

    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    ws = wb.create_sheet(title=sheet_name)
    bold = Font(bold=True)

    ws["A1"] = "Banks Master"
    ws["A1"].font = bold
    ws["A2"] = "Bank"
    ws["B2"] = "Address"
    ws["C2"] = "Latitude"
    ws["D2"] = "Longitude"
    for c in ["A2", "B2", "C2", "D2"]:
        ws[c].font = bold

    r = 3
    for bank, address in banks:
        ws.cell(row=r, column=1).value = bank
        ws.cell(row=r, column=2).value = address
        ws.cell(row=r, column=3).value = None
        ws.cell(row=r, column=4).value = None
        r += 1

    ws["F1"] = "Services Master"
    ws["F1"].font = bold
    ws["F2"] = "Service"
    ws["F2"].font = bold
    r = 3
    for svc in services:
        ws.cell(row=r, column=6).value = svc
        r += 1

    ws.freeze_panes = "A3"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ==========================================
# LOAD FILE (long cache + manual refresh)
# ==========================================
if st.button("Refresh data", key="bank_periodics_refresh_data"):
    download_banks_excel_cached.clear()
    get_visible_sheet_names.clear()
    detect_header_row.clear()
    read_sheet_with_detected_header.clear()
    st.rerun()

try:
    with st.spinner("Syncing banks data..."):
        local_path = download_banks_excel_cached(BANKS_SP_PATH)
except Exception as e:
    st.error(str(e))
    st.stop()

EXCEL_PATH_LOCAL = Path(local_path)
if not EXCEL_PATH_LOCAL.exists():
    st.error("Banks cache file missing after download.")
    st.stop()

visible_sheets = get_visible_sheet_names(str(EXCEL_PATH_LOCAL))
if not visible_sheets:
    st.error("No visible sheets found.")
    st.stop()

month_sheets = [s for s in visible_sheets if str(s).strip().lower() != "masters"]
if not month_sheets:
    st.error("No month sheets found.")
    st.stop()

views = ["Edit Status", "Report", "Report Matrix"]
if "bank_periodics_view" not in st.session_state:
    st.session_state["bank_periodics_view"] = "Edit Status"
elif st.session_state["bank_periodics_view"] not in views:
    st.session_state["bank_periodics_view"] = "Edit Status"

st.radio(
    "View",
    options=views,
    key="bank_periodics_view",
    horizontal=True,
)

st.markdown(
    """
    <style>
    div[data-testid="stRadio"] div[role="radiogroup"]{
        flex-wrap: nowrap;
        overflow-x: auto;
        overflow-y: hidden;
        scrollbar-width: thin;
        width: 100%;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] > label{
        flex: 0 0 auto;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

if st.session_state["bank_periodics_view"] == "Report":
    sheet = st.selectbox(
        "Month",
        options=month_sheets,
        key="report_sheet",
        index=default_last_month_index(month_sheets),
    )
    header_row = detect_header_row(str(EXCEL_PATH_LOCAL), sheet)
    df_raw = read_sheet_with_detected_header(str(EXCEL_PATH_LOCAL), sheet, header_row)

    if df_raw.empty:
        st.info("No data found on this sheet.")
        st.stop()

    bank_col = find_required_col(df_raw, BANK_FALLBACKS)
    addr_col = find_required_col(df_raw, ADDRESS_FALLBACKS)
    if not bank_col or not addr_col:
        st.error("Bank or Address column not found.")
        st.stop()

    df_raw[bank_col] = to_text_series(df_raw[bank_col])
    df_raw[addr_col] = to_text_series(df_raw[addr_col])

    task_cols = [c for c in df_raw.columns if c not in {bank_col, addr_col}]

    st.subheader("Map")
    masters_buildings = load_masters_buildings_table(str(EXCEL_PATH_LOCAL))
    map_df = map_dataset_from_left_join_by_location(
        masters_df=masters_buildings,
        month_df=df_raw,
        bank_col=bank_col,
        addr_col=addr_col,
    )
    if map_df.empty:
        st.info("Could not build map data from `Masters`.")
    else:
        status_options = ["Pending", "Not Pending"]
        selected_statuses = st.multiselect(
            "Status",
            options=status_options,
            default=status_options,
            key="report_map_status_filter",
        )
        if selected_statuses:
            map_df = map_df[map_df["Status"].isin(selected_statuses)]

        if map_df.empty:
            st.info("No map points for the selected status filter.")
            st.stop()

        map_df = map_df.copy()
        map_df["PendingText"] = map_df["PendingCount"].astype(int).astype(str)
        center_lat = float(map_df["Latitude"].mean())
        center_lon = float(map_df["Longitude"].mean())
        folium_map = folium.Map(
            location=[center_lat, center_lon],
            zoom_start=5,
            tiles="OpenStreetMap",
            control_scale=True,
            prefer_canvas=True,
        )
        cluster = MarkerCluster(
            name="Locations",
            options={"maxClusterRadius": 28},
            spiderfy_on_max_zoom=True,
            show_coverage_on_hover=False,
            zoom_to_bounds_on_click=True,
        ).add_to(folium_map)

        for _, row in map_df.iterrows():
            bank = str(row["Bank"])
            addr = str(row["Address"])
            cnt = int(row["PendingCount"])
            lat = float(row["Latitude"])
            lon = float(row["Longitude"])
            bank_color = BANK_STYLES.get(bank, {}).get("bg", "#1e88e5")
            text_color = _text_color_for_bg(bank_color)

            # Circle-style marker with black border + centered count.
            folium.Marker(
                location=[lat, lon],
                tooltip=f"Bank: {bank}<br>Address: {addr}<br>Pending Services: {cnt}",
                icon=DivIcon(
                    icon_size=(24, 24),
                    icon_anchor=(12, 12),
                    html=(
                        "<div style=\""
                        "width:24px;height:24px;border-radius:50%;"
                        "border:2px solid black;"
                        f"background:{bank_color};"
                        "display:flex;align-items:center;justify-content:center;"
                        f"font-size:12px;font-weight:700;color:{text_color};"
                        "line-height:1;"
                        "user-select:none;\">"
                        f"{cnt}"
                        "</div>"
                    ),
                ),
            ).add_to(cluster)

        st_folium(folium_map, width=None, height=700)

    st.subheader("Summary")
    bank_vals = sorted(df_raw[bank_col].dropna().unique().tolist())
    addr_vals = sorted(df_raw[addr_col].dropna().unique().tolist())

    c1, c2 = st.columns(2)
    with c1:
        bank_sel = st.multiselect("Filter: Bank", options=bank_vals, default=[])
    with c2:
        addr_sel = st.multiselect("Filter: Address", options=addr_vals, default=[])

    banks_to_use = bank_sel if bank_sel else bank_vals
    addrs_to_use = addr_sel if addr_sel else addr_vals
    df = df_raw[df_raw[bank_col].isin(banks_to_use) & df_raw[addr_col].isin(addrs_to_use)]
    done_pending_by_column_barchart(df, task_cols)

elif st.session_state["bank_periodics_view"] == "Report Matrix":
    st.subheader("Report Matrix")

    sheet_matrix = st.selectbox(
        "Month",
        options=month_sheets,
        key="matrix_sheet",
        index=default_last_month_index(month_sheets),
    )
    header_row_matrix = detect_header_row(str(EXCEL_PATH_LOCAL), sheet_matrix)
    df_matrix_raw = read_sheet_with_detected_header(str(EXCEL_PATH_LOCAL), sheet_matrix, header_row_matrix)

    if df_matrix_raw.empty:
        st.info("No data found on this sheet.")
        st.stop()

    bank_col_matrix = find_required_col(df_matrix_raw, BANK_FALLBACKS)
    addr_col_matrix = find_required_col(df_matrix_raw, ADDRESS_FALLBACKS)
    if not bank_col_matrix or not addr_col_matrix:
        st.error("Bank or Address column not found.")
        st.stop()

    df_matrix_raw[bank_col_matrix] = to_text_series(df_matrix_raw[bank_col_matrix])
    df_matrix_raw[addr_col_matrix] = to_text_series(df_matrix_raw[addr_col_matrix])

    bank_vals_matrix = sorted(df_matrix_raw[bank_col_matrix].dropna().unique().tolist())
    addr_vals_matrix = sorted(df_matrix_raw[addr_col_matrix].dropna().unique().tolist())

    c1, c2 = st.columns(2)
    with c1:
        bank_sel_matrix = st.multiselect("Filter: Bank", options=bank_vals_matrix, default=[], key="matrix_bank_filter")
    with c2:
        addr_sel_matrix = st.multiselect("Filter: Address", options=addr_vals_matrix, default=[], key="matrix_addr_filter")

    banks_to_use_matrix = bank_sel_matrix if bank_sel_matrix else bank_vals_matrix
    addrs_to_use_matrix = addr_sel_matrix if addr_sel_matrix else addr_vals_matrix
    df_matrix = df_matrix_raw[
        df_matrix_raw[bank_col_matrix].isin(banks_to_use_matrix)
        & df_matrix_raw[addr_col_matrix].isin(addrs_to_use_matrix)
    ]

    if df_matrix.empty:
        st.info("No rows match filters.")
    else:
        df_show = df_matrix.copy()
        for c in df_show.columns:
            df_show[c] = df_show[c].map(lambda v: None if _is_blank(v) else str(v).strip())
        st.dataframe(style_table(df_show, bank_col=bank_col_matrix, addr_col=addr_col_matrix), use_container_width=True, hide_index=True)

    st.subheader("Modify Record")
    task_cols_matrix = [c for c in df_matrix_raw.columns if c not in {bank_col_matrix, addr_col_matrix}]
    if not task_cols_matrix:
        st.info("No service columns found.")
        st.stop()

    bank_options_edit = sorted(df_matrix_raw[bank_col_matrix].dropna().unique().tolist())
    if not bank_options_edit:
        st.info("No banks available.")
        st.stop()

    selected_bank_matrix = st.selectbox("Bank", options=bank_options_edit, key="matrix_edit_bank")
    selected_service_matrix = st.selectbox("Service", options=task_cols_matrix, key="matrix_edit_service")

    rows_for_bank = df_matrix_raw[df_matrix_raw[bank_col_matrix] == selected_bank_matrix].copy()
    if rows_for_bank.empty:
        st.info("No rows found for selected bank.")
        st.stop()

    rows_for_bank["__address"] = rows_for_bank[addr_col_matrix].astype(str).str.strip()
    rows_for_bank["__status"] = rows_for_bank[selected_service_matrix].astype(str).str.strip()
    rows_for_bank["__label"] = rows_for_bank["__address"] + " | current: " + rows_for_bank["__status"]
    address_labels = rows_for_bank["__label"].tolist()
    selected_label = st.selectbox("Address", options=address_labels, key="matrix_edit_address")
    selected_row = rows_for_bank[rows_for_bank["__label"] == selected_label].iloc[0]
    selected_address_matrix = str(selected_row["__address"]).strip()
    current_status_matrix = selected_row[selected_service_matrix]

    st.caption(f"Current status: {current_status_matrix if not _is_blank(current_status_matrix) else '(empty)'}")
    new_status_matrix = st.selectbox(
        "New status",
        options=["Done"],
        index=0,
        key="matrix_edit_new_status",
    )

    if st.button("Save Matrix Record", type="primary", key="matrix_save_status_change"):
        try:
            token = get_token_silent_or_raise(
                "Not authenticated. Please connect in the main app (app.py).",
                "Session expired. Please reconnect in the main app (app.py).",
            )
            drive_id = resolve_drive_id(token)

            latest_bytes = download_sharepoint_file_bytes(BANKS_SP_PATH, token, drive_id=drive_id)
            updated_bytes = update_service_status_in_workbook(
                workbook_bytes=latest_bytes,
                sheet_name=sheet_matrix,
                bank_col=bank_col_matrix,
                addr_col=addr_col_matrix,
                service_col=selected_service_matrix,
                bank_value=selected_bank_matrix,
                address_value=selected_address_matrix,
                new_status=new_status_matrix,
            )
            upload_sharepoint_file_bytes(BANKS_SP_PATH, updated_bytes, token, drive_id=drive_id)

            download_banks_excel_cached.clear()
            get_visible_sheet_names.clear()
            detect_header_row.clear()
            read_sheet_with_detected_header.clear()

            st.success("Record updated in SharePoint Excel.")
            st.rerun()
        except Exception as e:
            st.error(f"Could not update status: {e}")

else:
    st.subheader("Update Bank Service Status")

    sheet_edit = st.selectbox(
        "Month",
        options=month_sheets,
        key="edit_sheet",
        index=default_last_month_index(month_sheets),
    )
    header_row_edit = detect_header_row(str(EXCEL_PATH_LOCAL), sheet_edit)
    df_edit = read_sheet_with_detected_header(str(EXCEL_PATH_LOCAL), sheet_edit, header_row_edit)

    if df_edit.empty:
        st.info("No data found on this sheet.")
    else:
        bank_col_edit = find_required_col(df_edit, BANK_FALLBACKS)
        addr_col_edit = find_required_col(df_edit, ADDRESS_FALLBACKS)

        if not bank_col_edit or not addr_col_edit:
            st.error("Bank or Address column not found.")
        else:
            df_edit[bank_col_edit] = to_text_series(df_edit[bank_col_edit])
            df_edit[addr_col_edit] = to_text_series(df_edit[addr_col_edit])

            task_cols_edit = [c for c in df_edit.columns if c not in {bank_col_edit, addr_col_edit}]
            pending_long = build_pending_long(df_edit, bank_col_edit, addr_col_edit, task_cols_edit)
            matrix = pending_matrix_by_bank_service(pending_long)

            st.markdown("**Pending Matrix (Bank x Service)**")
            if matrix.empty:
                st.info("No pending items found in this sheet.")
            else:
                st.dataframe(style_pending_matrix(matrix), use_container_width=True, hide_index=True)

            if not pending_long.empty:
                bank_options = sorted(pending_long["Bank"].unique().tolist())
                selected_bank = st.selectbox("Bank", options=bank_options, key="edit_bank")

                pending_bank = pending_long[pending_long["Bank"] == selected_bank]
                address_options = sorted(pending_bank["Address"].unique().tolist())
                selected_address = st.selectbox("Address", options=address_options, key="edit_address")

                pending_addr = pending_bank[pending_bank["Address"] == selected_address]
                service_options = sorted(pending_addr["Service"].unique().tolist())
                service_col = st.selectbox("Service (Pending only)", options=service_options, key="edit_service")

                st.caption("Current status: Pending")

                new_status = st.selectbox(
                    "New status",
                    options=["Done"],
                    index=0,
                    key="edit_new_status",
                )

                if st.button("Save Status Change", type="primary"):
                    try:
                        token = get_token_silent_or_raise(
                            "Not authenticated. Please connect in the main app (app.py).",
                            "Session expired. Please reconnect in the main app (app.py).",
                        )
                        drive_id = resolve_drive_id(token)

                        latest_bytes = download_sharepoint_file_bytes(BANKS_SP_PATH, token, drive_id=drive_id)
                        updated_bytes = update_service_status_in_workbook(
                            workbook_bytes=latest_bytes,
                            sheet_name=sheet_edit,
                            bank_col=bank_col_edit,
                            addr_col=addr_col_edit,
                            service_col=service_col,
                            bank_value=selected_bank,
                            address_value=selected_address,
                            new_status=new_status,
                        )
                        upload_sharepoint_file_bytes(BANKS_SP_PATH, updated_bytes, token, drive_id=drive_id)

                        download_banks_excel_cached.clear()
                        get_visible_sheet_names.clear()
                        detect_header_row.clear()
                        read_sheet_with_detected_header.clear()

                        st.success("Status updated in SharePoint Excel.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not update status: {e}")
